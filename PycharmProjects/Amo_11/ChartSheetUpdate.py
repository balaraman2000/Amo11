
import os
import sys
import pywintypes
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import GlobalLogs


class ChartSheetUpdate:
    def __init__(self, file_path):
        self.file_path = file_path


        if os.path.exists(self.file_path):  # Check if the file exists
            self.update_chart_sheet()

        else:
            #print(f"Workbook '{self.file_path}' not found.")
            GlobalLogs.logging.error(f"Workbook '{self.file_path}' not found.")

    def update_chart_sheet(self):
        # Load workbook using openpyxl to access its data
        try:
            wb = load_workbook(self.file_path)

            # Get the active sheet


            sheet = wb.active

            # Find the last data column
            last_data_column_index = None
            for column in range(1, sheet.max_column + 1):
                if sheet.cell(row=13, column=column).value is not None:
                    last_data_column_index = column

            # Calculate the last data cell
            if last_data_column_index is not None:
                last_data_column_letter = get_column_letter(last_data_column_index)
                self.las_num = f"{last_data_column_letter}"
                GlobalLogs.logging.info(f"Index of the last data cell in row 13:'{self.las_num}'")
            else:
                GlobalLogs.logging.error("No data found.")

            # Close the workbook
            wb.close()
        except ValueError as e:
            print("Error: Sheet name not found in the workbook")
            wb.close()
            sys.exit()
        except pywintypes.com_error as e:
            print(f"Error:Sheet not found")
            GlobalLogs.logging.error(f"Sheet not found{e}")
            wb.close()
            sys.exit()
        except Exception as e:
            # Print or log the error message for debugging purposes
            print(f"An error occurred.")
            GlobalLogs.logging.error(f"An error occurred.{e}")
            wb.close()
            sys.exit()


        try:
            # Try to open the workbook using xlwings
            wb = xw.Book(self.file_path)
        except ValueError:
            # If the workbook is already open, close it and try again
            for app in xw.apps:
                for book in app.books:
                    if book.name == os.path.basename(self.file_path):
                        book.close()
            wb = xw.Book(self.file_path)  # Try opening the workbook again

        # Access the chart sheet
        try:
            sheet = wb.sheets['DB容量']

            for chart in sheet.charts:
                GlobalLogs.logging.info(chart.name)

            target_chart_name = 'Chart 3'
            chart = None
            for c in sheet.charts:
                if c.name == target_chart_name:
                    chart = c
                    break

            if chart is not None:
                chart.set_source_data(sheet.range(f'DB容量!$C$13:${self.las_num}$31'))
                print("Graph Updated")
                ("Graph Updated")
                wb.save()
                wb.close()
            else:
                print(f"Error:Chart '{target_chart_name}' not found in DB容量 Sheet.")
                GlobalLogs.logging.error(f"Chart '{target_chart_name}' not found in DB容量 sheet.")
                wb.save()
                wb.close()
                sys.exit()
        except ValueError as e:
            print("Error: Sheet name not found in the workbook")
            wb.close()
            sys.exit()
        except pywintypes.com_error as e:
            print(f"Error:DB容量 Sheet not found")
            GlobalLogs.logging.error(f"Error:DB容量 Sheet not found")
            wb.close()
            sys.exit()
        except Exception as e:
            # Print or log the error message for debugging purposes
            print(f"An error occurred.")
            GlobalLogs.logging.error(f"An error occurred.{e}")
            wb.close()
            sys.exit()





